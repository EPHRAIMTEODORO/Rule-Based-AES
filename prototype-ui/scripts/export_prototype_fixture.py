from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
FEATURES_PATH = ROOT / "NewAes" / "ELAT_DATA" / "full_sample_aes_features_with_grammar_llm_improved.xlsx"
ESSAYS_PATH = ROOT / "NewAes" / "ELAT_DATA" / "essays.xlsx"
OUTPUT_PATH = ROOT / "prototype-ui" / "src" / "data" / "aesCases.json"


def rows_from_sheet(path: Path, sheet_name: str | None = None) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def round_float(value: object, places: int = 3) -> object:
    if isinstance(value, float):
        return round(value, places)
    return value


def main() -> None:
    essays = {row["essay_id"]: row for row in rows_from_sheet(ESSAYS_PATH, "essays")}
    cases = []

    for row in rows_from_sheet(FEATURES_PATH):
        essay = essays.get(row["essay_id"], {})
        source_score = row.get("score")
        llm_score = row.get("llm_overall_score")
        agreement_gap = (
            abs(float(llm_score) - float(source_score))
            if isinstance(source_score, (int, float)) and isinstance(llm_score, (int, float))
            else None
        )

        cases.append(
            {
                "essayId": row.get("essay_id"),
                "promptId": essay.get("prompt_id", "Unknown prompt"),
                "group": essay.get("Group", "Unknown"),
                "essayText": essay.get("text", ""),
                "sourceScore": round_float(source_score),
                "llmOverallScore": round_float(llm_score),
                "agreementGap": round_float(agreement_gap),
                "llmPerformanceBand": row.get("llm_performance_band"),
                "featureMeasures": {
                    "wordCount": row.get("word_count"),
                    "sentenceCount": row.get("sentence_count"),
                    "meanSentenceLength": round_float(row.get("mean_sentence_length")),
                    "avgWordLength": round_float(row.get("avg_word_length")),
                    "typeTokenRatio": round_float(row.get("type_token_ratio")),
                    "awlRatio": round_float(row.get("awl_ratio")),
                    "connectiveDensity": round_float(row.get("connective_density")),
                    "lexicalOverlap": round_float(row.get("lexical_overlap")),
                    "grammarErrors": row.get("grammar_errors"),
                    "grammarErrorsPer100": round_float(row.get("grammar_errors_per_100")),
                    "llmParagraphCount": row.get("llm_paragraph_count"),
                },
                "llmRubric": {
                    "organizationCoherence": row.get("llm_organization_coherence"),
                    "paragraphDevelopment": row.get("llm_paragraph_development"),
                    "supportingDetailElaboration": row.get("llm_supporting_detail_elaboration"),
                    "comprehensibility": row.get("llm_comprehensibility"),
                    "promptFulfillment": row.get("llm_prompt_fulfillment"),
                },
                "llmJustification": row.get("llm_justification", ""),
            }
        )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(cases, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(cases)} cases to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
