"""HITL hybrid AES pipeline using rule-based features plus local Llama 3.

This script intentionally does not modify the original rule-based extractor.
It imports `aes_feature_extractor.evaluate_essay`, computes paragraph count,
calls a local Ollama Llama 3 model, and writes a CSV with both `aes_*` and
`llm_*` columns plus the template-facing columns used by
`01_Raw_Inputs` in the HITL academic writing workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Optional


THIS_DIR = Path(__file__).resolve().parent
PROJECT_DIR = THIS_DIR.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from llm_rubric_prompt import SYSTEM_PROMPT, build_user_prompt  # noqa: E402


TEXT_COLUMN_CANDIDATES = [
    "essay",
    "text",
    "text_clean",
    "essay_text",
    "response",
    "answer",
]

FEATURE_NAMES = [
    "word_count",
    "sentence_count",
    "mean_sentence_length",
    "type_token_ratio",
    "avg_word_length",
    "avg_word_freq",
    "mtld",
    "awl_ratio",
    "clause_density",
    "dependency_depth",
    "noun_complexity",
    "connective_density",
    "lexical_overlap",
    "grammar_errors_per_100",
]

LLM_FIELD_NAMES = [
    "organization",
    "paragraph_development",
    "supporting_detail",
    "abstract_elaboration",
    "prompt_control",
    "comprehensibility",
    "grammar_meaning_impact",
    "recommended_score",
    "justification",
]

INTEGER_TRAIT_FIELDS = [
    "organization",
    "paragraph_development",
    "supporting_detail",
    "abstract_elaboration",
    "prompt_control",
    "comprehensibility",
    "grammar_meaning_impact",
]

HITL_RAW_INPUT_FIELDNAMES = [
    "Essay_ID",
    "Essay_Text",
    "LLM_Organization_Coherence",
    "LLM_Supporting_Detail",
    "LLM_Paragraph_Development",
    "LLM_Comprehensibility",
    "LLM_Prompt_Fulfillment",
    "Word_Count",
    "Sentence_Count",
    "Paragraph_Count",
    "Grammar_Errors_per_100",
    "Type_Token_Ratio",
    "Avg_Word_Length",
    "AWL_Ratio",
]

DECIMAL_PLACES = 5


def normalize_cell(value: object) -> str:
    """Convert input cells into clean strings."""
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\u00c2\u00a0": " ",
        "\u00a0": " ",
        "\u202f": " ",
        "\u2007": " ",
        "\ufeff": "",
        "¬†": " ",
    }
    for bad_value, replacement in replacements.items():
        text = text.replace(bad_value, replacement)
    return text.strip()


def choose_column(fieldnames: list[str], explicit: Optional[str], candidates: list[str]) -> str:
    """Choose a column from explicit input or known common names."""
    if explicit:
        if explicit not in fieldnames:
            raise ValueError(f"Input file does not contain column: {explicit}")
        return explicit

    lowered_to_original = {fieldname.lower(): fieldname for fieldname in fieldnames}
    for candidate in candidates:
        if candidate in lowered_to_original:
            return lowered_to_original[candidate]

    raise ValueError(
        "Could not infer text column. "
        f"Use --text-column with one of: {', '.join(fieldnames)}"
    )


def get_optional_column(row: dict, column_name: Optional[str], default: str = "") -> str:
    """Read an optional column from a row."""
    if not column_name:
        return default
    return normalize_cell(row.get(column_name, default))


def paragraph_count(text: str) -> int:
    """Estimate visible paragraph count from blank-line separated blocks."""
    cleaned = normalize_cell(text)
    if not cleaned:
        return 0

    blocks = [
        block.strip()
        for block in re.split(r"(?:\r?\n\s*){2,}", cleaned)
        if block.strip()
    ]
    if blocks:
        return len(blocks)

    lines = [line.strip() for line in cleaned.splitlines() if line.strip()]
    return len(lines) if lines else 1


def alphabetic_words(text: str) -> list[str]:
    """Return lowercase alphabetic words for template lexical fields."""
    return re.findall(r"\b[a-zA-Z]+\b", text.lower())


def safe_divide(numerator: float, denominator: float, default: float = 0.0) -> float:
    """Divide safely and return a default when the denominator is zero."""
    if denominator == 0:
        return default
    return numerator / denominator


def round_decimal(value: float) -> float:
    """Round numeric feature values consistently with the base AES extractor."""
    return round(value, DECIMAL_PLACES)


def extract_hitl_lexical_features(text: str) -> dict:
    """Compute lexical fields required by the HITL workbook template."""
    words = alphabetic_words(text)
    return {
        "type_token_ratio": round_decimal(safe_divide(len(set(words)), len(words))),
        "avg_word_length": round_decimal(
            safe_divide(sum(len(word) for word in words), len(words))
        ),
    }


def clamp_int_score(value: object, min_value: int = 1, max_value: int = 6) -> int:
    """Convert an LLM trait score to a bounded integer."""
    try:
        numeric = int(round(float(value)))
    except (TypeError, ValueError):
        return min_value
    return max(min_value, min(max_value, numeric))


def clamp_half_point_score(value: object, min_value: float = 1.0, max_value: float = 6.0) -> float:
    """Convert an LLM recommended score to a bounded half-point value."""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return min_value
    numeric = max(min_value, min(max_value, numeric))
    return round(numeric * 2) / 2


def extract_json_object(text: str) -> dict:
    """Parse a JSON object, allowing for accidental text around it."""
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not match:
        raise ValueError("LLM response did not contain a JSON object.")

    return json.loads(match.group(0))


def ollama_health_url(ollama_url: str) -> str:
    """Convert an Ollama chat API URL into a lightweight health-check URL."""
    parsed_url = urllib.parse.urlparse(ollama_url)
    return urllib.parse.urlunparse(parsed_url._replace(path="/api/tags", query=""))


def is_ollama_running(ollama_url: str, timeout_seconds: float = 2.0) -> bool:
    """Return whether the Ollama API is already reachable."""
    request = urllib.request.Request(ollama_health_url(ollama_url), method="GET")
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds):
            return True
    except (urllib.error.URLError, TimeoutError, OSError):
        return False


def start_ollama_if_needed(args: argparse.Namespace) -> None:
    """Start `ollama serve` when the configured Ollama API is not reachable."""
    if is_ollama_running(args.ollama_url):
        if not args.quiet:
            print("Ollama is already running.", flush=True)
        return

    if not args.quiet:
        print("Starting Ollama server...", flush=True)

    try:
        process = subprocess.Popen(
            [args.ollama_command, "serve"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except FileNotFoundError as exc:
        raise RuntimeError(
            "Could not find the Ollama command. Install Ollama or pass "
            "--ollama-command with the full path to the executable."
        ) from exc

    deadline = time.monotonic() + args.ollama_startup_timeout
    while time.monotonic() < deadline:
        if is_ollama_running(args.ollama_url):
            if not args.quiet:
                print("Ollama server is ready.", flush=True)
            return
        if process.poll() is not None:
            break
        time.sleep(0.5)

    raise RuntimeError(
        "Tried to start Ollama, but the API did not become reachable. "
        f"Check that `{args.ollama_command} serve` works and that "
        f"{ollama_health_url(args.ollama_url)} is accessible."
    )


def call_ollama_chat(
    system_prompt: str,
    user_prompt: str,
    model: str,
    ollama_url: str,
    temperature: float,
    timeout_seconds: int,
) -> dict:
    """Call Ollama's local chat API and return the parsed JSON response."""
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "format": "json",
        "stream": False,
        "options": {
            "temperature": temperature,
        },
    }
    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        ollama_url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(
            "Ollama returned an HTTP error. Make sure the model name matches "
            f"`ollama list`. Requested model: {model}. "
            f"HTTP {exc.code}: {error_body}"
        ) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(
            "Could not reach Ollama. Make sure Ollama is running and the model is "
            f"available with: ollama run {model}"
        ) from exc

    content = response_payload.get("message", {}).get("content", "")
    if not content:
        raise ValueError(f"Ollama returned no message content: {response_payload}")

    return extract_json_object(content)


def normalize_llm_result(raw_result: dict, essay_id: str) -> dict:
    """Normalize LLM JSON into stable output columns."""
    result = {"essay_id": essay_id}

    for field_name in INTEGER_TRAIT_FIELDS:
        result[field_name] = clamp_int_score(raw_result.get(field_name))

    result["recommended_score"] = clamp_half_point_score(
        raw_result.get(
            "llm_recommended_score",
            raw_result.get("recommended_score", raw_result.get("overall_score")),
        )
    )
    result["justification"] = normalize_cell(raw_result.get("justification", ""))
    return result


def flatten_aes_result(result: dict, paragraph_total: int, essay_text: str) -> dict:
    """Flatten the rule-based AES result into CSV columns."""
    features = {
        **result["features"],
        **extract_hitl_lexical_features(essay_text),
    }
    row = {
        "aes_score": result["score"],
        "aes_paragraph_count": paragraph_total,
    }
    for feature_name in FEATURE_NAMES:
        row[f"aes_{feature_name}"] = features.get(feature_name)
    return row


def flatten_llm_result(result: dict) -> dict:
    """Flatten the LLM rubric result into CSV columns."""
    return {
        f"llm_{field_name}": result.get(field_name, "")
        for field_name in LLM_FIELD_NAMES
    }


def flatten_hitl_raw_input_aliases(
    essay_id: str,
    essay_text: str,
    aes_row: dict,
    llm_result: dict,
) -> dict:
    """Build workbook-ready aliases for the HITL template raw-input sheet."""
    return {
        "Essay_ID": essay_id,
        "Essay_Text": essay_text,
        "LLM_Organization_Coherence": llm_result.get("organization", ""),
        "LLM_Supporting_Detail": llm_result.get("supporting_detail", ""),
        "LLM_Paragraph_Development": llm_result.get("paragraph_development", ""),
        "LLM_Comprehensibility": llm_result.get("comprehensibility", ""),
        "LLM_Prompt_Fulfillment": llm_result.get("prompt_control", ""),
        "Word_Count": aes_row.get("aes_word_count", ""),
        "Sentence_Count": aes_row.get("aes_sentence_count", ""),
        "Paragraph_Count": aes_row.get("aes_paragraph_count", ""),
        "Grammar_Errors_per_100": aes_row.get("aes_grammar_errors_per_100", ""),
        "Type_Token_Ratio": aes_row.get("aes_type_token_ratio", ""),
        "Avg_Word_Length": aes_row.get("aes_avg_word_length", ""),
        "AWL_Ratio": aes_row.get("aes_awl_ratio", ""),
    }


def read_csv_rows(input_path: str) -> tuple[list[str], list[dict]]:
    """Read rows from a CSV file."""
    with open(input_path, newline="", encoding="utf-8-sig") as input_file:
        reader = csv.DictReader(input_file)
        if reader.fieldnames is None:
            raise ValueError("Input CSV must include a header row.")
        return list(reader.fieldnames), list(reader)


def read_xlsx_rows(input_path: str, sheet_name: Optional[str]) -> tuple[list[str], list[dict]]:
    """Read rows from an XLSX file."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading .xlsx files requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Input XLSX sheet is empty.") from exc

    fieldnames = [
        normalize_cell(value) or f"column_{index}"
        for index, value in enumerate(header_row, start=1)
    ]

    output_rows = []
    for row_values in rows:
        if all(value is None for value in row_values):
            continue
        row = {
            fieldname: normalize_cell(value)
            for fieldname, value in zip(fieldnames, row_values)
        }
        for fieldname in fieldnames[len(row_values) :]:
            row[fieldname] = ""
        output_rows.append(row)

    workbook.close()
    return fieldnames, output_rows


def read_input_rows(input_path: str, sheet_name: Optional[str]) -> tuple[list[str], list[dict]]:
    """Read input rows from CSV or XLSX."""
    suffix = Path(input_path).suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(input_path)
    if suffix == ".xlsx":
        return read_xlsx_rows(input_path, sheet_name)
    raise ValueError("Input file must be a .csv or .xlsx file.")


def write_csv_rows(output_path: str, fieldnames: list[str], rows: list[dict]) -> None:
    """Write merged rows to a CSV file."""
    with open(output_path, "w", newline="", encoding="utf-8") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def evaluate_row(row: dict, args: argparse.Namespace, essay_index: int, text_column: str) -> dict:
    """Evaluate one essay with rule-based features and LLM features."""
    from aes_feature_extractor import evaluate_essay

    essay_text = normalize_cell(row.get(text_column, ""))
    essay_id = get_optional_column(row, args.essay_id_column, f"essay_{essay_index}")
    prompt_id = get_optional_column(row, args.prompt_column, "Not provided")
    paragraph_total = paragraph_count(essay_text)

    aes_result = evaluate_essay(essay_text)
    llm_prompt = build_user_prompt(
        essay_id=essay_id,
        prompt_id=prompt_id,
        paragraph_count=paragraph_total,
        essay_text=essay_text,
    )
    raw_llm_result = call_ollama_chat(
        system_prompt=SYSTEM_PROMPT,
        user_prompt=llm_prompt,
        model=args.model,
        ollama_url=args.ollama_url,
        temperature=args.temperature,
        timeout_seconds=args.timeout,
    )
    llm_result = normalize_llm_result(raw_llm_result, essay_id)
    aes_row = flatten_aes_result(aes_result, paragraph_total, essay_text)

    return {
        **{fieldname: normalize_cell(value) for fieldname, value in row.items()},
        **aes_row,
        **flatten_llm_result(llm_result),
        **flatten_hitl_raw_input_aliases(
            essay_id=essay_id,
            essay_text=essay_text,
            aes_row=aes_row,
            llm_result=llm_result,
        ),
    }


def build_output_fieldnames(original_fieldnames: list[str]) -> list[str]:
    """Build stable output field names."""
    generated_fieldnames = [
        "aes_score",
        "aes_paragraph_count",
        *[f"aes_{feature_name}" for feature_name in FEATURE_NAMES],
        *[f"llm_{field_name}" for field_name in LLM_FIELD_NAMES],
        *HITL_RAW_INPUT_FIELDNAMES,
    ]
    return original_fieldnames + [
        fieldname
        for fieldname in generated_fieldnames
        if fieldname not in original_fieldnames
    ]


def evaluate_file(args: argparse.Namespace) -> None:
    """Evaluate an input file and write hybrid AES results."""
    if not args.no_start_ollama:
        start_ollama_if_needed(args)

    original_fieldnames, input_rows = read_input_rows(args.input, args.sheet)
    text_column = choose_column(
        original_fieldnames,
        args.text_column,
        TEXT_COLUMN_CANDIDATES,
    )
    rows_to_process = input_rows[: args.limit] if args.limit else input_rows

    output_rows = []
    for essay_index, row in enumerate(rows_to_process, start=1):
        if not args.quiet:
            print(f"Scoring essay {essay_index}/{len(rows_to_process)}", flush=True)
        output_rows.append(evaluate_row(row, args, essay_index, text_column))
        if args.delay_seconds:
            time.sleep(args.delay_seconds)

    write_csv_rows(
        args.output,
        build_output_fieldnames(original_fieldnames),
        output_rows,
    )


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate hybrid AES features using rule-based features and local Llama 3."
    )
    parser.add_argument("--input", "-i", required=True, help="Input CSV or XLSX file.")
    parser.add_argument("--output", "-o", required=True, help="Output CSV file.")
    parser.add_argument("--text-column", help="Column containing essay text.")
    parser.add_argument("--essay-id-column", help="Column containing essay IDs.")
    parser.add_argument("--prompt-column", help="Column containing prompt/topic IDs.")
    parser.add_argument("--sheet", help="Optional XLSX sheet name.")
    parser.add_argument("--model", default="llama3:8b", help="Ollama model name.")
    parser.add_argument(
        "--ollama-url",
        default="http://127.0.0.1:11434/api/chat",
        help="Ollama chat API URL.",
    )
    parser.add_argument(
        "--ollama-command",
        default="ollama",
        help="Ollama executable to use when auto-starting the server.",
    )
    parser.add_argument(
        "--ollama-startup-timeout",
        type=float,
        default=30.0,
        help="Seconds to wait for `ollama serve` to become reachable.",
    )
    parser.add_argument(
        "--no-start-ollama",
        action="store_true",
        help="Do not auto-start Ollama before scoring.",
    )
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--limit", type=int, help="Only process the first N essays.")
    parser.add_argument(
        "--delay-seconds",
        type=float,
        default=0.0,
        help="Optional delay between LLM calls.",
    )
    parser.add_argument("--quiet", action="store_true", help="Suppress progress output.")
    return parser.parse_args()


if __name__ == "__main__":
    evaluate_file(parse_args())
