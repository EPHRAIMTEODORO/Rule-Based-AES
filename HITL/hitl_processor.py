"""Workbook processor for the HITL desktop-app backend.

The desktop app should call `process_workbook` with an uploaded Excel file and
an output path. The function validates the upload, scores each essay with the
existing HITL hybrid scorer, writes a completed `.xlsx`, and returns structured
records that the UI can render.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


THIS_DIR = Path(__file__).resolve().parent
PROJECT_DIR = THIS_DIR.parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

import hybrid_llm_aes as scorer  # noqa: E402


DEFAULT_ID_COLUMN = "id"
DEFAULT_PROMPT_COLUMN = "Topic"
DEFAULT_TEXT_COLUMN = "Essay"
DEFAULT_MODEL = "llama3:8b"
DEFAULT_OLLAMA_URL = "http://127.0.0.1:11434/api/chat"
HUMAN_DECISION_FIELDNAMES = [
    "Rater_Final_Score",
    "Rater_Final_Placement",
    "Rater_Action",
    "Decision_Status",
    "Admissions_Decision",
    "Reason_Notes",
    "Decision_Updated_At",
]

ProgressCallback = Callable[[dict], None]


@dataclass
class ProcessingResult:
    """Structured result returned to the desktop UI."""

    input_path: str
    output_path: str
    rows_processed: int
    columns: list[str]
    records: list[dict]

    def to_dict(self) -> dict:
        """Return a JSON-serializable representation."""
        return asdict(self)


@dataclass
class ProcessorOptions:
    """Configuration used while processing an uploaded workbook."""

    essay_id_column: Optional[str] = DEFAULT_ID_COLUMN
    prompt_column: Optional[str] = DEFAULT_PROMPT_COLUMN
    text_column: Optional[str] = DEFAULT_TEXT_COLUMN
    sheet_name: Optional[str] = None
    model: str = DEFAULT_MODEL
    ollama_url: str = DEFAULT_OLLAMA_URL
    ollama_command: str = "ollama"
    ollama_startup_timeout: float = 30.0
    start_ollama: bool = True
    temperature: float = 0.0
    timeout: int = 180
    limit: Optional[int] = None
    delay_seconds: float = 0.0
    quiet: bool = True


def _emit_progress(callback: Optional[ProgressCallback], **payload: object) -> None:
    """Send a progress event to the UI when a callback is available."""
    if callback:
        callback(dict(payload))


def _is_blank_header(value: object) -> bool:
    """Return whether a header cell should be ignored."""
    return scorer.normalize_cell(value) == ""


def _read_workbook_rows(
    input_path: str,
    sheet_name: Optional[str] = None,
) -> tuple[list[str], list[dict]]:
    """Read an uploaded workbook and drop blank header columns."""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        workbook.close()
        raise ValueError("Uploaded workbook sheet is empty.") from exc

    kept_columns = [
        (index, scorer.normalize_cell(value))
        for index, value in enumerate(header_row)
        if not _is_blank_header(value)
    ]
    if not kept_columns:
        workbook.close()
        raise ValueError("Uploaded workbook must include a header row.")

    fieldnames = [fieldname for _, fieldname in kept_columns]
    output_rows = []
    for row_values in rows:
        if all(value is None for value in row_values):
            continue

        row = {}
        for source_index, fieldname in kept_columns:
            value = row_values[source_index] if source_index < len(row_values) else ""
            row[fieldname] = scorer.normalize_cell(value)
        output_rows.append(row)

    workbook.close()
    return fieldnames, output_rows


def _validate_upload_columns(
    fieldnames: list[str],
    options: ProcessorOptions,
) -> str:
    """Validate upload columns and return the selected essay text column."""
    text_column = scorer.choose_column(
        fieldnames,
        options.text_column,
        scorer.TEXT_COLUMN_CANDIDATES,
    )

    for label, column_name in (
        ("essay ID", options.essay_id_column),
        ("prompt/topic", options.prompt_column),
    ):
        if column_name and column_name not in fieldnames:
            raise ValueError(
                f"Uploaded workbook does not contain the {label} column: {column_name}"
            )

    return text_column


def _build_scorer_args(options: ProcessorOptions) -> argparse.Namespace:
    """Build the small namespace expected by the existing row scorer."""
    return argparse.Namespace(
        essay_id_column=options.essay_id_column,
        prompt_column=options.prompt_column,
        model=options.model,
        ollama_url=options.ollama_url,
        ollama_command=options.ollama_command,
        ollama_startup_timeout=options.ollama_startup_timeout,
        temperature=options.temperature,
        timeout=options.timeout,
        delay_seconds=options.delay_seconds,
        quiet=options.quiet,
        no_start_ollama=not options.start_ollama,
    )


def _autosize_columns(worksheet) -> None:
    """Set readable column widths without letting essays make huge columns."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in column_cells[1:]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        if header in {"Essay", "Essay_Text", "llm_justification"}:
            worksheet.column_dimensions[column_letter].width = 48
        else:
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 24)


def _write_completed_workbook(
    output_path: str,
    fieldnames: list[str],
    rows: list[dict],
) -> None:
    """Write scored rows to a completed Excel workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Completed_Scores"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, fieldname in enumerate(fieldnames, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=fieldname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, fieldname in enumerate(fieldnames, start=1):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=row.get(fieldname, ""),
            )
            cell.alignment = Alignment(wrap_text=fieldname in {"Essay", "Essay_Text", "llm_justification"})

    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def write_completed_workbook(
    output_path: str,
    fieldnames: list[str],
    rows: list[dict],
) -> None:
    """Write scored and rater-reviewed rows to a completed Excel workbook."""
    _write_completed_workbook(output_path, fieldnames, rows)


def _with_initial_decision_fields(row: dict) -> dict:
    """Add default human decision fields to a scored row."""
    return {
        **row,
        "Rater_Final_Score": row.get("Rater_Final_Score", ""),
        "Rater_Final_Placement": row.get("Rater_Final_Placement", ""),
        "Rater_Action": row.get("Rater_Action", ""),
        "Decision_Status": row.get("Decision_Status") or "Pending",
        "Admissions_Decision": row.get("Admissions_Decision", ""),
        "Reason_Notes": row.get("Reason_Notes", ""),
        "Decision_Updated_At": row.get("Decision_Updated_At", ""),
    }


def process_workbook(
    input_path: str,
    output_path: str,
    essay_id_column: Optional[str] = DEFAULT_ID_COLUMN,
    prompt_column: Optional[str] = DEFAULT_PROMPT_COLUMN,
    text_column: Optional[str] = DEFAULT_TEXT_COLUMN,
    sheet_name: Optional[str] = None,
    model: str = DEFAULT_MODEL,
    ollama_url: str = DEFAULT_OLLAMA_URL,
    ollama_command: str = "ollama",
    ollama_startup_timeout: float = 30.0,
    start_ollama: bool = True,
    temperature: float = 0.0,
    timeout: int = 180,
    limit: Optional[int] = None,
    delay_seconds: float = 0.0,
    quiet: bool = True,
    progress_callback: Optional[ProgressCallback] = None,
) -> ProcessingResult:
    """Process an uploaded workbook and write a completed workbook.

    The returned records are the same rows written to the Excel output, making
    them suitable for immediate display in the desktop UI.
    """
    options = ProcessorOptions(
        essay_id_column=essay_id_column,
        prompt_column=prompt_column,
        text_column=text_column,
        sheet_name=sheet_name,
        model=model,
        ollama_url=ollama_url,
        ollama_command=ollama_command,
        ollama_startup_timeout=ollama_startup_timeout,
        start_ollama=start_ollama,
        temperature=temperature,
        timeout=timeout,
        limit=limit,
        delay_seconds=delay_seconds,
        quiet=quiet,
    )
    scorer_args = _build_scorer_args(options)

    _emit_progress(progress_callback, stage="starting", message="Preparing workbook")
    original_fieldnames, input_rows = _read_workbook_rows(input_path, sheet_name)
    selected_text_column = _validate_upload_columns(original_fieldnames, options)
    rows_to_process = input_rows[:limit] if limit else input_rows

    if not rows_to_process:
        raise ValueError("Uploaded workbook does not contain any essay rows.")

    if options.start_ollama:
        _emit_progress(progress_callback, stage="ollama", message="Checking Ollama")
        scorer.start_ollama_if_needed(scorer_args)

    output_rows = []
    total = len(rows_to_process)
    for essay_index, row in enumerate(rows_to_process, start=1):
        _emit_progress(
            progress_callback,
            stage="scoring",
            current=essay_index,
            total=total,
            message=f"Scoring essay {essay_index} of {total}",
        )
        output_rows.append(
            _with_initial_decision_fields(
                scorer.evaluate_row(row, scorer_args, essay_index, selected_text_column)
            )
        )
        if delay_seconds:
            scorer.time.sleep(delay_seconds)

    output_fieldnames = scorer.build_output_fieldnames(original_fieldnames) + [
        fieldname
        for fieldname in HUMAN_DECISION_FIELDNAMES
        if fieldname not in original_fieldnames
    ]
    _emit_progress(progress_callback, stage="writing", message="Writing completed workbook")
    _write_completed_workbook(output_path, output_fieldnames, output_rows)

    _emit_progress(
        progress_callback,
        stage="complete",
        rows_processed=len(output_rows),
        output_path=str(output_path),
        message="Completed workbook is ready",
    )
    return ProcessingResult(
        input_path=str(input_path),
        output_path=str(output_path),
        rows_processed=len(output_rows),
        columns=output_fieldnames,
        records=output_rows,
    )


def _parse_args() -> argparse.Namespace:
    """Parse command-line arguments for backend testing."""
    parser = argparse.ArgumentParser(
        description="Process an uploaded HITL essay workbook into a completed XLSX."
    )
    parser.add_argument("--input", "-i", required=True, help="Uploaded workbook path.")
    parser.add_argument("--output", "-o", required=True, help="Completed workbook path.")
    parser.add_argument("--essay-id-column", default=DEFAULT_ID_COLUMN)
    parser.add_argument("--prompt-column", default=DEFAULT_PROMPT_COLUMN)
    parser.add_argument("--text-column", default=DEFAULT_TEXT_COLUMN)
    parser.add_argument("--sheet", dest="sheet_name")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--ollama-url", default=DEFAULT_OLLAMA_URL)
    parser.add_argument("--ollama-command", default="ollama")
    parser.add_argument("--ollama-startup-timeout", type=float, default=30.0)
    parser.add_argument("--no-start-ollama", action="store_true")
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--delay-seconds", type=float, default=0.0)
    parser.add_argument("--quiet", action="store_true")
    return parser.parse_args()


def _print_progress(event: dict) -> None:
    """Print simple CLI progress events."""
    message = event.get("message")
    if message:
        print(message, flush=True)


def main() -> None:
    """Run the processor as a command-line smoke-test utility."""
    args = _parse_args()
    result = process_workbook(
        input_path=args.input,
        output_path=args.output,
        essay_id_column=args.essay_id_column,
        prompt_column=args.prompt_column,
        text_column=args.text_column,
        sheet_name=args.sheet_name,
        model=args.model,
        ollama_url=args.ollama_url,
        ollama_command=args.ollama_command,
        ollama_startup_timeout=args.ollama_startup_timeout,
        start_ollama=not args.no_start_ollama,
        temperature=args.temperature,
        timeout=args.timeout,
        limit=args.limit,
        delay_seconds=args.delay_seconds,
        quiet=args.quiet,
        progress_callback=None if args.quiet else _print_progress,
    )
    if args.quiet:
        print(result.to_dict(), flush=True)


if __name__ == "__main__":
    main()
