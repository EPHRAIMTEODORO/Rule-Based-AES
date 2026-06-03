"""Rule-based Automated Essay Scoring feature extractor.

This module extracts interpretable lexical, syntactic, cohesion, and grammar
features from an essay and combines a small subset into a transparent weighted
score. It does not use machine learning or LLM-based scoring.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from pathlib import Path
from typing import Optional

if os.environ.get("JAVA_HOME"):
    os.environ["PATH"] = (
        str(Path(os.environ["JAVA_HOME"]) / "bin")
        + os.pathsep
        + os.environ.get("PATH", "")
    )

import language_tool_python
import spacy
from lexical_diversity import lex_div as ld
from nltk.tokenize import sent_tokenize
from spacy.tokens import Doc, Token
from wordfreq import zipf_frequency


# Load heavyweight NLP resources once at module import time.
nlp = spacy.load("en_core_web_sm")

try:
    tool: Optional[language_tool_python.LanguageTool] = (
        language_tool_python.LanguageTool("en-US")
    )
except Exception:
    tool = None


FALLBACK_AWL_SET = {
    "analyze",
    "approach",
    "area",
    "assess",
    "assume",
    "authority",
    "available",
    "benefit",
    "concept",
    "consistent",
    "context",
    "data",
    "derive",
    "distribute",
    "economy",
    "environment",
    "establish",
    "estimate",
    "evidence",
    "factor",
    "function",
    "identify",
    "indicate",
    "interpret",
    "involve",
    "issue",
    "method",
    "occur",
    "percent",
    "period",
    "policy",
    "principle",
    "proceed",
    "process",
    "require",
    "research",
    "respond",
    "section",
    "significant",
    "similar",
    "source",
    "specific",
    "structure",
    "theory",
    "variable",
}

CLAUSE_DEPENDENCIES = {"ccomp", "xcomp", "advcl"}

CONNECTIVES = {
    "however",
    "therefore",
    "because",
    "although",
    "moreover",
    "thus",
    "also",
    "furthermore",
    "consequently",
    "nevertheless",
    "in addition",
    "for example",
    "for instance",
    "on the other hand",
    "as a result",
}

FEATURE_NAMES = [
    "word_count",
    "sentence_count",
    "mean_sentence_length",
    "avg_word_freq",
    "mtld",
    "awl_ratio",
    "clause_density",
    "dependency_depth",
    "noun_complexity",
    "connective_density",
    "lexical_overlap",
    "grammar_errors_per_100",
]

TEXT_COLUMN_CANDIDATES = [
    "essay",
    "text",
    "text_clean",
    "essay_text",
    "response",
    "answer",
]
RESULT_PREFIX = "aes_"
DECIMAL_PLACES = 5
PROJECT_DIR = Path(__file__).resolve().parent
AWL_DATA_PATH = PROJECT_DIR / "data" / "awl_word_forms.json"


def _load_awl_set(path: Path = AWL_DATA_PATH) -> set[str]:
    """Load full AWL word forms from JSON, falling back to the starter list."""
    try:
        with path.open(encoding="utf-8") as file:
            rows = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return FALLBACK_AWL_SET

    words = {
        str(row.get("word", "")).lower()
        for row in rows
        if isinstance(row, dict) and row.get("word")
    }
    return words or FALLBACK_AWL_SET


AWL_SET = _load_awl_set()


def safe_divide(numerator: float, denominator: float, default: float = 0.0) -> float:
    """Divide safely and return a default when the denominator is zero."""
    if denominator == 0:
        return default
    return numerator / denominator


def dependency_depth(token: Token) -> int:
    """Return the number of dependency hops from a token to the sentence root."""
    depth = 0
    while token.head != token:
        token = token.head
        depth += 1
    return depth


def _alphabetic_words(doc: Doc) -> list[str]:
    """Return lowercased alphabetic spaCy tokens."""
    return [token.text.lower() for token in doc if token.is_alpha]


def _sentence_count(doc: Doc, text: str) -> int:
    """Count sentences from spaCy, with a fallback for very short/raw input."""
    if not text.strip():
        return 0

    sentences = [sent for sent in doc.sents if sent.text.strip()]
    return len(sentences) if sentences else 1


def _average_word_frequency(words: list[str]) -> float:
    """Average Zipf word frequency; lower values suggest rarer vocabulary."""
    frequencies = [zipf_frequency(word, "en") for word in words]
    return safe_divide(sum(frequencies), len(frequencies))


def _mtld(words: list[str]) -> float:
    """Measure of Textual Lexical Diversity, guarded for short essays."""
    if len(words) < 2:
        return 0.0

    try:
        return float(ld.mtld(words))
    except Exception:
        return 0.0


def _awl_ratio(words: list[str]) -> float:
    """Ratio of words that appear in the full PDF-derived AWL word-form set."""
    awl_count = sum(1 for word in words if word in AWL_SET)
    return safe_divide(awl_count, len(words))


def _clause_density(doc: Doc, word_count: int) -> float:
    """Approximate clause density using selected dependency labels."""
    clause_count = sum(1 for token in doc if token.dep_ in CLAUSE_DEPENDENCIES)
    return safe_divide(clause_count, word_count)


def _average_dependency_depth(doc: Doc) -> float:
    """Average dependency tree depth across non-space tokens."""
    tokens = [token for token in doc if not token.is_space]
    depths = [dependency_depth(token) for token in tokens]
    return safe_divide(sum(depths), len(depths))


def _noun_complexity(doc: Doc) -> float:
    """Average number of syntactic children attached to each noun token."""
    nouns = [token for token in doc if token.pos_ == "NOUN"]
    child_counts = [len(list(noun.children)) for noun in nouns]
    return safe_divide(sum(child_counts), len(child_counts))


def _connective_density(text: str, word_count: int) -> float:
    """Count single- and multi-word connectives per 100 words."""
    lowered_text = text.lower()
    connective_count = 0

    for connective in CONNECTIVES:
        pattern = rf"\b{re.escape(connective)}\b"
        connective_count += len(re.findall(pattern, lowered_text))

    return safe_divide(connective_count * 100, word_count)


def _safe_sent_tokenize(text: str) -> list[str]:
    """Use NLTK sentence tokenization with a simple fallback if data is missing."""
    try:
        return sent_tokenize(text)
    except LookupError:
        return [
            part.strip()
            for part in re.split(r"(?<=[.!?])\s+", text)
            if part.strip()
        ]


def _word_set(sentence: str) -> set[str]:
    """Return a lowercase alphabetic word set for lexical overlap."""
    return set(re.findall(r"\b[a-zA-Z]+\b", sentence.lower()))


def _lexical_overlap(text: str) -> float:
    """Average Jaccard overlap between adjacent sentence word sets."""
    sentences = _safe_sent_tokenize(text)
    if len(sentences) < 2:
        return 0.0

    overlaps = []
    for first, second in zip(sentences, sentences[1:]):
        first_words = _word_set(first)
        second_words = _word_set(second)
        union = first_words | second_words
        overlaps.append(safe_divide(len(first_words & second_words), len(union)))

    return safe_divide(sum(overlaps), len(overlaps))


def _grammar_errors_per_100(text: str, word_count: int) -> Optional[float]:
    """LanguageTool grammar matches per 100 words, or None if checking fails."""
    if word_count == 0:
        return 0.0

    if tool is None:
        return None

    try:
        matches = tool.check(text)
    except Exception:
        return None

    return safe_divide(len(matches) * 100, word_count)


def extract_features(text: str) -> dict:
    """Extract interpretable rule-based AES features from essay text."""
    cleaned_text = text or ""
    doc = nlp(cleaned_text)
    words = _alphabetic_words(doc)

    word_count = len(words)
    sentence_count = _sentence_count(doc, cleaned_text)

    return {
        "word_count": word_count,
        "sentence_count": sentence_count,
        "mean_sentence_length": safe_divide(word_count, sentence_count),
        "avg_word_freq": _average_word_frequency(words),
        "mtld": _mtld(words),
        "awl_ratio": _awl_ratio(words),
        "clause_density": _clause_density(doc, word_count),
        "dependency_depth": _average_dependency_depth(doc),
        "noun_complexity": _noun_complexity(doc),
        "connective_density": _connective_density(cleaned_text, word_count),
        "lexical_overlap": _lexical_overlap(cleaned_text),
        "grammar_errors_per_100": _grammar_errors_per_100(cleaned_text, word_count),
    }


def _round_result_value(value: object) -> object:
    """Round floating-point result values for cleaner reporting."""
    if isinstance(value, float):
        return round(value, DECIMAL_PLACES)
    return value


def _round_feature_values(features: dict) -> dict:
    """Round numeric feature values without changing feature names."""
    return {
        feature_name: _round_result_value(value)
        for feature_name, value in features.items()
    }


def evaluate_essays(texts: list[str]) -> list[dict]:
    """Evaluate many essays and preserve their original input order."""
    return [
        {
            "essay_index": index,
            **evaluate_essay(text),
        }
        for index, text in enumerate(texts, start=1)
    ]


def normalize(
    value: Optional[float],
    min_val: float,
    max_val: float,
    invert: bool = False,
) -> float:
    """Normalize a value to 0-1, optionally inverting so lower values score higher."""
    if value is None or max_val == min_val:
        return 0.0

    normalized = (value - min_val) / (max_val - min_val)
    normalized = max(0.0, min(1.0, normalized))

    if invert:
        return 1.0 - normalized
    return normalized


def compute_score(features: dict) -> float:
    """Compute a transparent weighted score from selected features."""
    score = (
        normalize(features.get("mtld"), 20, 100) * 0.25
        + normalize(features.get("awl_ratio"), 0, 0.20) * 0.30
        + normalize(features.get("avg_word_freq"), 2, 6, invert=True) * 0.20
        + normalize(features.get("grammar_errors_per_100"), 0, 20, invert=True) * 0.25
    )
    return round(score * 100, DECIMAL_PLACES)


def evaluate_essay(text: str) -> dict:
    """Return the weighted score and full feature dictionary for an essay."""
    features = extract_features(text)
    return {
        "score": compute_score(features),
        "features": _round_feature_values(features),
    }


def _result_fieldnames(prefix: str = RESULT_PREFIX) -> list[str]:
    """Return CSV column names for generated AES results."""
    return [
        f"{prefix}essay_index",
        f"{prefix}score",
        *[f"{prefix}{feature_name}" for feature_name in FEATURE_NAMES],
    ]


def _flatten_result(
    result: dict,
    essay_index: int,
    prefix: str = RESULT_PREFIX,
) -> dict:
    """Flatten a nested evaluation result into one CSV-friendly row."""
    row = {
        f"{prefix}essay_index": essay_index,
        f"{prefix}score": result["score"],
    }

    for feature_name in FEATURE_NAMES:
        row[f"{prefix}{feature_name}"] = result["features"].get(feature_name)

    return row


def _choose_text_column(
    fieldnames: list[str],
    text_column: Optional[str] = None,
) -> str:
    """Choose the essay-text column, using common names when not specified."""
    if text_column is not None:
        if text_column not in fieldnames:
            raise ValueError(f"Input file does not contain text column: {text_column}")
        return text_column

    lowered_to_original = {fieldname.lower(): fieldname for fieldname in fieldnames}
    for candidate in TEXT_COLUMN_CANDIDATES:
        if candidate in lowered_to_original:
            return lowered_to_original[candidate]

    raise ValueError(
        "Could not infer the essay text column. "
        f"Use --text-column with one of: {', '.join(fieldnames)}"
    )


def _merge_original_with_result(
    original_row: dict,
    essay_index: int,
    text_column: str,
) -> dict:
    """Evaluate one input row and return original columns plus AES columns."""
    cleaned_row = {
        fieldname: _cell_to_csv_value(value)
        for fieldname, value in original_row.items()
    }
    essay_text = cleaned_row.get(text_column, "") or ""
    result = evaluate_essay(str(essay_text))
    return {
        **cleaned_row,
        **_flatten_result(result, essay_index),
    }


def _write_rows_to_csv(
    output_path: str,
    original_fieldnames: list[str],
    output_rows: list[dict],
) -> None:
    """Write original columns followed by generated AES columns to a CSV file."""
    fieldnames = original_fieldnames + [
        name for name in _result_fieldnames() if name not in original_fieldnames
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(
            {
                fieldname: _csv_output_value(row.get(fieldname, ""))
                for fieldname in fieldnames
            }
            for row in output_rows
        )


def _csv_output_value(value: object) -> object:
    """Format CSV values consistently while keeping non-floats unchanged."""
    if isinstance(value, float):
        return f"{value:.{DECIMAL_PLACES}f}"
    return value


def evaluate_essays_from_csv(
    input_path: str,
    output_path: str,
    text_column: Optional[str] = None,
) -> None:
    """Evaluate essays from a CSV file and preserve all original columns."""
    with open(input_path, newline="", encoding="utf-8-sig") as input_file:
        reader = csv.DictReader(input_file)

        if reader.fieldnames is None:
            raise ValueError("Input CSV must include a header row.")

        original_fieldnames = list(reader.fieldnames)
        selected_text_column = _choose_text_column(original_fieldnames, text_column)

        output_rows = [
            _merge_original_with_result(row, essay_index, selected_text_column)
            for essay_index, row in enumerate(reader, start=1)
        ]

    _write_rows_to_csv(output_path, original_fieldnames, output_rows)


def _cell_to_csv_value(value: object) -> object:
    """Convert Excel cell values into CSV-friendly values."""
    if value is None:
        return ""
    if isinstance(value, str):
        return _clean_spacing_artifacts(value)
    return value


def _clean_spacing_artifacts(text: str) -> str:
    """Convert hidden or mojibake non-breaking spaces into regular spaces."""
    replacements = {
        "\u00c2\u00a0": " ",
        "\u00a0": " ",
        "\u202f": " ",
        "\u2007": " ",
        "\ufeff": "",
        "¬†": " ",
    }

    cleaned_text = text
    for bad_value, replacement in replacements.items():
        cleaned_text = cleaned_text.replace(bad_value, replacement)

    return cleaned_text


def evaluate_essays_from_xlsx(
    input_path: str,
    output_path: str,
    text_column: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> None:
    """Evaluate essays from an XLSX sheet and preserve all original columns."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading .xlsx files requires openpyxl. "
            "Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Input XLSX sheet is empty.") from exc

    original_fieldnames = [
        str(value).strip() if value is not None else f"column_{index}"
        for index, value in enumerate(header_row, start=1)
    ]
    selected_text_column = _choose_text_column(original_fieldnames, text_column)

    output_rows = []
    for essay_index, row_values in enumerate(rows, start=1):
        if all(value is None for value in row_values):
            continue

        original_row = {
            fieldname: _cell_to_csv_value(value)
            for fieldname, value in zip(original_fieldnames, row_values)
        }

        for fieldname in original_fieldnames[len(row_values) :]:
            original_row[fieldname] = ""

        output_rows.append(
            _merge_original_with_result(original_row, essay_index, selected_text_column)
        )

    workbook.close()
    _write_rows_to_csv(output_path, original_fieldnames, output_rows)


def evaluate_essays_from_file(
    input_path: str,
    output_path: str,
    text_column: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> None:
    """Evaluate essays from a CSV or XLSX file and write a CSV result file."""
    suffix = Path(input_path).suffix.lower()

    if suffix == ".csv":
        evaluate_essays_from_csv(input_path, output_path, text_column)
    elif suffix == ".xlsx":
        evaluate_essays_from_xlsx(input_path, output_path, text_column, sheet_name)
    else:
        raise ValueError("Input file must be a .csv or .xlsx file.")


def _parse_args() -> argparse.Namespace:
    """Parse optional command-line arguments for batch file processing."""
    parser = argparse.ArgumentParser(
        description="Extract rule-based AES features from one essay, CSV, or XLSX file."
    )
    parser.add_argument(
        "--input",
        "-i",
        help="Path to a CSV or XLSX file containing essays.",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Path where the scored CSV should be written.",
    )
    parser.add_argument(
        "--text-column",
        help="Column containing essay text. If omitted, common names are inferred.",
    )
    parser.add_argument(
        "--sheet",
        help="Optional XLSX sheet name. Defaults to the first sheet.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()

    if args.input:
        if not args.output:
            raise SystemExit("Please provide --output when using --input.")

        evaluate_essays_from_file(
            input_path=args.input,
            output_path=args.output,
            text_column=args.text_column,
            sheet_name=args.sheet,
        )
        print(f"Wrote batch results to {args.output}")
    else:
        sample = (
            "Students benefit from evidence-based writing instruction. "
            "However, effective assessment also requires consistent methods, "
            "because teachers need specific data to identify patterns in student work."
        )
        result = evaluate_essay(sample)
        print(result)
